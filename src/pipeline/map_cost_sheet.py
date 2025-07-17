import pandas as pd
import os
import sys
from openpyxl import load_workbook
from pathlib import Path

# Support both script and PyInstaller .exe paths
if getattr(sys, 'frozen', False):
    # Running as PyInstaller executable
    SCRIPT_DIR = Path(sys._MEIPASS)
    OMNI_TEMPLATE_PATH = SCRIPT_DIR / "Files" / "OCTF-1539-COST SHEET.xlsx"
else:
    # Running as script - Files folder is in root directory, go up two levels from src/pipeline/
    SCRIPT_DIR = Path(__file__).parent
    OMNI_TEMPLATE_PATH = SCRIPT_DIR.parent.parent / "Files" / "OCTF-1539-COST SHEET.xlsx"

def generate_output_path(input_file_path, suffix, output_directory=None, extension=".xlsx"):
    """Generate output file path based on input file and optional output directory."""
    input_path = Path(input_file_path)
    base_name = input_path.stem
    output_filename = f"{base_name}{suffix}{extension}"
    
    if output_directory and output_directory.strip():
        output_dir = Path(output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_filename
    else:
        output_path = input_path.parent / output_filename
    
    return str(output_path)

def get_oem_file():
    while True:
        path = input("Enter path to OEMSecrets Excel file (with prices): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid file path.")

def get_merged_file():
    while True:
        path = input("Enter path to merged Excel file (with DESCRIPTION column): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid merged file path.")

def map_and_insert_data(oem_path, merged_path, template_path=OMNI_TEMPLATE_PATH):
    df_oem = pd.read_excel(oem_path, keep_default_na=False, na_values=[''])
    df_merged = pd.read_excel(merged_path, keep_default_na=False, na_values=[''])

    # Get the user's selected company from environment variable
    user_company = os.environ.get("BOM_COMPANY", "").lower()
    print(f"📊 User selected company: {user_company}")

    # Pull description column if present
    if "DESCRIPTION" in df_merged.columns:
        df_oem["DESCRIPTION"] = df_merged["DESCRIPTION"]
    elif "Description" in df_merged.columns:  # NEL format
        df_oem["DESCRIPTION"] = df_merged["Description"]
    else:
        print("⚠️ DESCRIPTION column not found in merged file. Skipping description.")
        df_oem["DESCRIPTION"] = ""

    # Pull PROTON P/N column from merged file for customer part number (NEL only)
    if user_company == "nel":
        if "Proton P/N" in df_merged.columns:
            df_oem["PROTON P/N"] = df_merged["Proton P/N"]
            print("✅ Added PROTON P/N column from merged file (NEL format)")
        else:
            print("⚠️ PROTON P/N column not found in merged file. Skipping customer part number.")
            df_oem["PROTON P/N"] = ""
    else:
        print("📊 Skipping PROTON P/N mapping (not NEL format)")
        df_oem["PROTON P/N"] = ""

    # Pull additional columns from merged file for Primetals format
    if user_company == "primetals":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Primetals format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Primetals format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MFG (manufacturer) from merged file
        if "MFG" in df_merged.columns:
            df_oem["MFG"] = df_merged["MFG"]
            print("✅ Added MFG column from merged file (Primetals format)")
        else:
            print("⚠️ MFG column not found in merged file. Skipping manufacturer.")
            df_oem["MFG"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Primetals format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for Riley Power format
    if user_company == "riley power":
        # Pull MANUFACTURER from merged file
        if "MANUFACTURER" in df_merged.columns:
            df_oem["MANUFACTURER"] = df_merged["MANUFACTURER"]
            print("✅ Added MANUFACTURER column from merged file (Riley Power format)")
        else:
            print("⚠️ MANUFACTURER column not found in merged file. Skipping manufacturer.")
            df_oem["MANUFACTURER"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Riley Power format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Riley Power format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Riley Power format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for Shanklin format
    if user_company == "shanklin":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Shanklin format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Shanklin format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Shanklin format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for 901D format
    if user_company == "901d":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (901D format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (901D format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MFR (manufacturer) from merged file
        if "MFR" in df_merged.columns:
            df_oem["MFR"] = df_merged["MFR"]
            print("✅ Added MFR column from merged file (901D format)")
        else:
            print("⚠️ MFR column not found in merged file. Skipping manufacturer.")
            df_oem["MFR"] = ""
        
        # Pull FIND NO. (item number) from merged file
        if "FIND NO." in df_merged.columns:
            df_oem["FIND NO."] = df_merged["FIND NO."]
            print("✅ Added FIND NO. column from merged file (901D format)")
        else:
            print("⚠️ FIND NO. column not found in merged file. Skipping item number.")
            df_oem["FIND NO."] = ""
        
        # Pull 901D P/N (customer part number) from merged file
        if "901D P/N" in df_merged.columns:
            df_oem["901D P/N"] = df_merged["901D P/N"]
            print("✅ Added 901D P/N column from merged file (901D format)")
        else:
            print("⚠️ 901D P/N column not found in merged file. Skipping customer part number.")
            df_oem["901D P/N"] = ""

    # Pull additional columns from merged file for Amazon format
    if user_company == "amazon":
        # Pull Part number from merged file
        if "Part number" in df_merged.columns:
            df_oem["Part number"] = df_merged["Part number"]
            print("✅ Added Part number column from merged file (Amazon format)")
        else:
            print("⚠️ Part number column not found in merged file. Skipping part number.")
            df_oem["Part number"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Amazon format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull Manufacturer from merged file
        if "Manufacturer" in df_merged.columns:
            df_oem["Manufacturer"] = df_merged["Manufacturer"]
            print("✅ Added Manufacturer column from merged file (Amazon format)")
        else:
            print("⚠️ Manufacturer column not found in merged file. Skipping manufacturer.")
            df_oem["Manufacturer"] = ""
        
        # Pull Device tag (item number) from merged file
        if "Device tag" in df_merged.columns:
            df_oem["Device tag"] = df_merged["Device tag"]
            print("✅ Added Device tag column from merged file (Amazon format)")
        else:
            print("⚠️ Device tag column not found in merged file. Skipping item number.")
            df_oem["Device tag"] = ""
        
        # Pull Distributor from merged file for supplier notes
        if "Distributor" in df_merged.columns:
            df_oem["Distributor"] = df_merged["Distributor"]
            print("✅ Added Distributor column from merged file (Amazon format)")
        else:
            print("⚠️ Distributor column not found in merged file. Skipping distributor.")
            df_oem["Distributor"] = ""

    # Detect company format based on column names
    oem_columns = df_oem.columns.tolist()
    merged_columns = df_merged.columns.tolist()
    
    print(f"📊 OEM file columns: {oem_columns}")
    print(f"📊 Merged file columns: {merged_columns}")
    
    # Debug: Check for price column variations
    price_columns = [col for col in oem_columns if 'price' in col.lower() or 'cost' in col.lower() or 'usd' in col.lower()]
    print(f"💰 Price-related columns found: {price_columns}")
    
    # Use user's selected company for mapping instead of auto-detection
    if user_company == "nel":
        print("📊 Using NEL format - based on user selection")
        column_mapping = {
            "Internal Reference": "OMNI PART #",
            "Part Number": "COMMERCIAL PART#",
            "Quantity for Single BOM": "UNIT QTY",
            "Extended Quantity for 1 BOM": "EXT QTY",
            "Manufacturer": "MFR",
            "Distributor": "SUPPLIER / NOTES",
            "Minimum Order": "MIN ORDER",
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "PROTON P/N": "CUST PART #",  # Pull from merged file for NEL
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "primetals":
        print("📊 Using Primetals format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MFG": "MFR",
            "MPN": "COMMERCIAL PART#",
            "QTY": "UNIT QTY",
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "riley power":
        print("📊 Using Riley Power format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MANUFACTURER": "MFR",  # Use MANUFACTURER from merged file
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file (converted from Part Number)
            "QTY": "UNIT QTY",  # Use QTY from merged file (converted from Quantity for Single BOM)
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "shanklin":
        print("📊 Using Shanklin format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file
            "QTY": "UNIT QTY",  # Use QTY from merged file
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "901d":
        print("📊 Using 901D format - based on user selection")
        column_mapping = {
            "FIND NO.": "ITEM",  # Use FIND NO. from merged file as item number
            "901D P/N": "CUST PART #",  # Use 901D P/N from merged file as customer part number
            "MFR": "MFR",  # Use MFR from merged file
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file
            "QTY": "UNIT QTY",  # Use QTY from merged file  
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",  # Alternative price column name
            "Price": "COST EACH",  # Alternative price column name
            "Cost": "COST EACH",  # Alternative price column name
            "Distributor": "SUPPLIER / NOTES",  # Map distributor to supplier field
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "amazon":
        print("📊 Using Amazon format - based on user selection")
        column_mapping = {
            "Device tag": "ITEM",  # Use Device tag from merged file as item number
            "QTY": "UNIT QTY",  # Use QTY from merged file
            "Manufacturer": "MFR",  # Use Manufacturer from merged file
            "Part number": "COMMERCIAL PART#",  # Use Part number from merged file
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",  # Alternative price column name
            "Price": "COST EACH",  # Alternative price column name
            "Cost": "COST EACH",  # Alternative price column name
            "Distributor": "SUPPLIER / NOTES",  # Map distributor to supplier field
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    else:
        print("📊 Using Farrell format - based on user selection")
        # Map the actual column names from the Farrell file
        column_mapping = {
            "Item": "ITEM",
            "Manufacturer": "MFR",
            "MPN": "COMMERCIAL PART#",
            "Quantity": "UNIT QTY",
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",  # Alternative price column name
            "Price": "COST EACH",  # Alternative price column name
            "Cost": "COST EACH",  # Alternative price column name
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }

    print(f"📊 Using column mapping: {column_mapping}")

    # Debug: Show which columns are actually being mapped
    columns_found = {k: v for k, v in column_mapping.items() if k in df_oem.columns}
    columns_missing = {k: v for k, v in column_mapping.items() if k not in df_oem.columns}
    print(f"✅ Columns found for mapping: {columns_found}")
    print(f"❌ Columns missing from OEM file: {columns_missing}")

    # Create a clean mapping that avoids duplicates - take only the first source for each target
    clean_mapping = {}
    for source_col, target_col in column_mapping.items():
        if source_col in df_oem.columns:
            if target_col not in clean_mapping.values():
                clean_mapping[source_col] = target_col
            else:
                print(f"⚠️ Skipping duplicate mapping: {source_col} -> {target_col} (already mapped)")
    
    print(f"📊 Clean mapping (no duplicates): {clean_mapping}")
    
    # Apply the clean mapping
    df_renamed = df_oem.rename(columns=clean_mapping)
    
    # Get all the target column names we want to keep
    mapped_target_columns = list(clean_mapping.values())
    df_out = df_renamed[mapped_target_columns].copy()
    
    print(f"📊 df_out shape after selection: {df_out.shape}")
    print(f"📊 df_out columns: {list(df_out.columns)}")

    # Handle missing values carefully - preserve numeric columns
    for col in df_out.columns:
        if col == "COST EACH":
            # For cost column, preserve numeric values but replace true NaN with 0
            print(f"💰 Processing COST EACH column...")
            print(f"💰 COST EACH sample values: {df_out[col].head().tolist()}")
            df_out[col] = pd.to_numeric(df_out[col], errors='coerce').fillna(0)
            print(f"💰 COST EACH after processing: {df_out[col].tolist()}")
        else:
            # For other columns, replace NaN with "N/A"
            df_out[col] = df_out[col].fillna("N/A")
            df_out[col] = df_out[col].replace("", "N/A")

    # Add ITEM numbers (sequential numbering)
    if "ITEM" not in df_out.columns:
        df_out["ITEM"] = range(1, len(df_out) + 1)

    print(f"📊 Final data shape: {df_out.shape}")
    print(f"📊 Final columns: {df_out.columns.tolist()}")
    if len(df_out) > 0:
        print(f"📊 Sample data:\n{df_out.head(2)}")
        # Debug: Show actual COST EACH values
        if "COST EACH" in df_out.columns:
            print(f"💰 COST EACH values: {df_out['COST EACH'].tolist()}")
            print(f"💰 COST EACH data types: {df_out['COST EACH'].dtype}")

    # Validate template file exists before trying to load it
    if not template_path.exists():
        raise FileNotFoundError(
            f"Cost sheet template not found: {template_path}\n\n"
            f"Please ensure the file structure is correct:\n"
            f"BoMination/\n"
            f"  Files/\n"
            f"    OCTF-1539-COST SHEET.xlsx\n"
            f"  src/\n"
            f"    (script files)"
        )

    print(f"LOADING: Loading cost sheet template: {template_path.name}")
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = 12
    excel_headers = {}
    for idx, cell in enumerate(ws[header_row]):
        val = cell.value
        if isinstance(val, str):
            # Clean up headers by normalizing spaces and converting to uppercase
            header = val.strip().upper().replace('\n', ' ').replace('\xa0', ' ')
            # Replace multiple spaces with single spaces
            import re
            header = re.sub(r'\s+', ' ', header)
            excel_headers[header] = idx + 1

    print("Mapped headers from cost sheet:", excel_headers)

    # Insert data row by row
    for i, row in df_out.iterrows():
        for col_name, value in row.items():
            col_upper = col_name.upper().strip()
            if col_upper in excel_headers:
                col_idx = excel_headers[col_upper]
                # Handle different data types appropriately
                if pd.notna(value) and str(value).strip() != '':
                    ws.cell(row=header_row + 1 + i, column=col_idx, value=value)
                    print(f"📝 Inserted '{value}' into {col_upper} at row {header_row + 1 + i}, col {col_idx}")
                else:
                    # Insert empty string for missing/null values
                    ws.cell(row=header_row + 1 + i, column=col_idx, value="")
            else:
                print(f"⚠️ Column '{col_upper}' not found in cost sheet template")

    print(f"✅ Inserted {len(df_out)} rows of data into cost sheet")

    # Save to PDF directory instead of using output_directory parameter
    # Get the original PDF path from the oem_path (which is based on the merged file)
    oem_path_obj = Path(oem_path)
    # Extract the base name without the suffixes to get back to the original PDF name
    base_name = oem_path_obj.stem.replace("_merged_with_prices", "").replace("_merged", "")
    cost_sheet_path = oem_path_obj.parent / f"{base_name}_cost_sheet.xlsx"
    
    print(f"📁 Saving cost sheet to: {cost_sheet_path}")
    wb.save(cost_sheet_path)
    print(f"Final cost sheet saved to: {cost_sheet_path}")

def main():
    """Main function to be called by the pipeline."""
    oem_path = os.environ.get("OEM_INPUT_PATH")
    merged_path = os.environ.get("MERGED_BOM_PATH")
    
    print(f"🐛 DEBUG: Environment variables:")
    print(f"  OEM_INPUT_PATH: {oem_path}")
    print(f"  MERGED_BOM_PATH: {merged_path}")
    print(f"  OEM file exists: {os.path.exists(oem_path) if oem_path else False}")
    print(f"  Merged file exists: {os.path.exists(merged_path) if merged_path else False}")
    
    if oem_path and os.path.exists(oem_path):
        # Quick check of what's in the OEM file
        try:
            import pandas as pd
            df_test = pd.read_excel(oem_path)
            print(f"🐛 DEBUG: OEM file '{oem_path}' has {df_test.shape[0]} rows and {df_test.shape[1]} columns")
            print(f"🐛 DEBUG: OEM file columns: {list(df_test.columns)}")
        except Exception as e:
            print(f"🐛 DEBUG: Error reading OEM file: {e}")
    
    map_and_insert_data(oem_path, merged_path)

if __name__ == "__main__":
    main()