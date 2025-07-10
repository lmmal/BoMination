import pandas as pd
import os
import sys
from openpyxl import load_workbook
from pathlib import Path

# Support both script and PyInstaller .exe paths
if getattr(sys, 'frozen', False):
    # Running as PyInstaller executable
    SCRIPT_DIR = Path(sys._MEIPASS)
    OMNI_TEMPLATE_PATH = SCRIPT_DIR / "Files" / "OCTF-1539-COST SHEET.xlsx"
else:
    # Running as script - Files folder is in parent directory of src
    SCRIPT_DIR = Path(__file__).parent
    OMNI_TEMPLATE_PATH = SCRIPT_DIR.parent / "Files" / "OCTF-1539-COST SHEET.xlsx"

def generate_output_path(input_file_path, suffix, output_directory=None, extension=".xlsx"):
    """Generate output file path based on input file and optional output directory."""
    input_path = Path(input_file_path)
    base_name = input_path.stem
    output_filename = f"{base_name}{suffix}{extension}"
    
    if output_directory and output_directory.strip():
        output_dir = Path(output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_filename
    else:
        output_path = input_path.parent / output_filename
    
    return str(output_path)

def get_oem_file():
    while True:
        path = input("Enter path to OEMSecrets Excel file (with prices): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid file path.")

def get_merged_file():
    while True:
        path = input("Enter path to merged Excel file (with DESCRIPTION column): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid merged file path.")

def map_and_insert_data(oem_path, merged_path, template_path=OMNI_TEMPLATE_PATH):
    df_oem = pd.read_excel(oem_path)
    df_merged = pd.read_excel(merged_path)

    # Pull description column if present
    if "DESCRIPTION" in df_merged.columns:
        df_oem["DESCRIPTION"] = df_merged["DESCRIPTION"]
    else:
        print("⚠️ DESCRIPTION column not found in merged file. Skipping description.")
        df_oem["DESCRIPTION"] = ""

    column_mapping = {
        "Internal Reference": "CUST PART #",
        "Manufacturer": "MFR",
        "Part Number": "COMMERCIAL PART#",
        "Quantity for Single BOM": "UNIT QTY",
        "Unit Price in USD": "COST EACH",
        "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
        "Distributor": "SUPPLIER / NOTES",
        "DESCRIPTION": "DESCRIPTION"
    }

    df_renamed = df_oem.rename(columns={k: v for k, v in column_mapping.items() if k in df_oem.columns})
    mapped_columns = list(column_mapping.values())
    df_out = df_renamed[[col for col in mapped_columns if col in df_renamed.columns]]

    # Validate template file exists before trying to load it
    if not template_path.exists():
        raise FileNotFoundError(
            f"Cost sheet template not found: {template_path}\n\n"
            f"Please ensure the file structure is correct:\n"
            f"BoMination/\n"
            f"  Files/\n"
            f"    OCTF-1539-COST SHEET.xlsx\n"
            f"  src/\n"
            f"    (script files)"
        )

    print(f"LOADING: Loading cost sheet template: {template_path.name}")
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = 12
    excel_headers = {}
    for idx, cell in enumerate(ws[header_row]):
        val = cell.value
        if isinstance(val, str):
            header = val.strip().upper().replace('\n', ' ').replace('\xa0', ' ')
            excel_headers[header] = idx + 1

    print("Mapped headers from cost sheet:", excel_headers)

    for i, row in df_out.iterrows():
        for col_name, value in row.items():
            col_upper = col_name.upper().strip()
            if col_upper in excel_headers:
                col_idx = excel_headers[col_upper]
                ws.cell(row=header_row + 1 + i, column=col_idx, value=value)

    # Save to PDF directory instead of using output_directory parameter
    # Get the original PDF path from the oem_path (which is based on the merged file)
    oem_path_obj = Path(oem_path)
    # Extract the base name without the suffixes to get back to the original PDF name
    base_name = oem_path_obj.stem.replace("_merged_with_prices", "").replace("_merged", "")
    cost_sheet_path = oem_path_obj.parent / f"{base_name}_cost_sheet.xlsx"
    
    print(f"📁 Saving cost sheet to: {cost_sheet_path}")
    wb.save(cost_sheet_path)
    print(f"Final cost sheet saved to: {cost_sheet_path}")

def main():
    """Main function to be called by the pipeline."""
    oem_path = os.environ.get("OEM_INPUT_PATH")
    merged_path = os.environ.get("MERGED_BOM_PATH")
    map_and_insert_data(oem_path, merged_path)

if __name__ == "__main__":
    main()